#!/usr/bin/env python3
"""把 tests/regression_test.py 的回归检查项导出为 Excel 清单

用法：
  python tools/export_regression_excel.py
产出：
  outputs/回归测试项目列表.xlsx
"""
import inspect
import os
import sys

# 确保 imports 能找到 tests
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.join(ROOT, "tests"))

import regression_test as rt

# 大类映射（函数 -> 中文）
CATEGORY_MAP = {
    "t_files": "文件完整性",
    "t_syntax": "语法门禁",
    "t_menu": "菜单黄金基线",
    "t_nav_links": "导航外链基线",
    "t_domains": "域名黄金基线",
    "t_apis": "接口端点基线",
    "t_globals": "全局变量基线",
    "t_styles": "join-us 关键样式基线",
    "t_webconfig": "IIS web.config 基线",
    "t_deploy": "部署文件基线",
    "t_src_site": "源码工程基线",
    "t_visual_checklist": "视觉对照 checklist",
    "t_online": "线上路由可达性",
}

rows = []


def capture_check(category, name, cond, detail=""):
    # 按冒号拆分为 小类 / 名称
    if ":" in name:
        sub, item = name.split(":", 1)
        sub = sub.strip()
        item = item.strip()
    else:
        sub, item = "默认", name.strip()
    rows.append({
        "大类": category,
        "小类": sub,
        "检查项名称": item,
        "说明/期望": str(detail) if detail else "",
    })


def main():
    # 备份原 check 函数
    original_check = rt.check

    # 扫描并依次调用每个 t_ 函数
    for fname in CATEGORY_MAP:
        func = getattr(rt, fname)

        def make_wrapper(cat):
            def wrapper(name, cond, detail=""):
                capture_check(cat, name, cond, detail)
            return wrapper

        rt.check = make_wrapper(CATEGORY_MAP[fname])
        try:
            func()
        except Exception as e:
            capture_check(CATEGORY_MAP[fname], f"[{fname} 执行异常]", False, str(e))

    rt.check = original_check

    # 写 Excel
    out_dir = os.path.join(ROOT, "outputs")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "回归测试项目列表.xlsx")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "回归测试项目"
    headers = ["序号", "大类", "小类", "检查项名称", "说明/期望"]
    ws.append(headers)

    for idx, r in enumerate(rows, 1):
        ws.append([idx, r["大类"], r["小类"], r["检查项名称"], r["说明/期望"]])

    # 样式
    header_fill = PatternFill("solid", fgColor="1A5319")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 40

    wb.save(out_path)
    print(f"✅ 已导出 {len(rows)} 条回归检查项: {out_path}")


if __name__ == "__main__":
    main()
